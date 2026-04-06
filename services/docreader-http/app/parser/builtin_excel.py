"""
builtin_excel：xlsx 用 openpyxl 只读遍历；xls 不做本地回退（依赖 MarkItDown 或报错）。
"""

from __future__ import annotations

import io
import logging
from typing import Any

from app.models.document import Document
from app.parser.base_parser import BaseParser

logger = logging.getLogger(__name__)


class ExcelBuiltinParser(BaseParser):
    """ExcelBuiltinParser 将表格按行展开为易检索的纯文本。"""

    def parse_into_text(self, content: bytes) -> Document:
        meta: dict[str, Any] = {"parser": "openpyxl"}
        if (self.file_type or "").lower() == "xls":
            return Document(metadata={**meta, "hint": "xls_no_local_fallback"})
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            return Document(metadata={**meta, "error": f"openpyxl:{e}"})

        lines: list[str] = []
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            for sheet in wb:
                lines.append(f"## {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None and str(c).strip() != ""]
                    if cells:
                        lines.append(" | ".join(cells))
            wb.close()
        except Exception as e:
            logger.exception("openpyxl failed")
            return Document(metadata={**meta, "error": str(e)})

        text = "\n\n".join(lines)
        if not text.strip():
            return Document(metadata=meta)
        return Document(content=text, metadata=meta)
